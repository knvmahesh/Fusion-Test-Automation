from mcp.server.fastmcp import FastMCP
from dotenv import load_dotenv
import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

load_dotenv("../.env")

# Create an MCP server
mcp = FastMCP(name="Test_Automation",host="0.0.0.0",port=8050)

#mcp = FastMCP(name="Test_Automation",port=8000)

#To get row count#
def getRowCount(file: str, sheetName: str) -> int:
    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = wb[sheetName]
    count = sheet.max_row
    wb.close()
    return count

#To get Column count#
def getColumnCount(file: str, sheetName: str) -> int:
    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = wb[sheetName]
    count = sheet.max_column
    wb.close()
    return count

#To Read Data from Excel#
def readData(file: str, sheetName: str, rownum: int, columnno: int):
    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = wb[sheetName]
    value = sheet.cell(row=rownum, column=columnno).value
    wb.close()
    return value

#To write Data to Excel#
def writeData(file: str, sheetName: str, rownum: int, columno: int, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    sheet.cell(row=rownum, column=columno).value = data
    wb.save(file)
    wb.close()

# Add a test_Automation tool for Oracle Fusion Login
@mcp.tool()
def Fusion_login(excel_path: str = r"C:\Users\velur\Desktop\Selenium-Python\Termination.xlsx"):
    """Fusion Login"""
    #driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    options = webdriver.ChromeOptions()
    options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    driver = webdriver.Chrome(
    service=ChromeService(ChromeDriverManager().install()), 
    options=options
    
    )
    try:
        driver.maximize_window()

        rows_URL = getRowCount(excel_path, 'URL')
        rows_DETAILS = getRowCount(excel_path, 'Termination')

        Base_url = readData(excel_path, "URL", 2, 1)
        driver.get(Base_url)

        UN = readData(excel_path, "LD", 2, 1)
        PW = readData(excel_path, "LD", 2, 2)

        driver.find_element(By.ID, "userid").send_keys(UN)
        driver.find_element(By.ID, "password").send_keys(PW)
        driver.find_element(By.ID, "btnActive").click()

        return {"status": "logged_in", "rows_URL": rows_URL, "rows_DETAILS": rows_DETAILS}
    finally:
        time.sleep(2)
        try:
            driver.quit()
        except Exception:
            pass

# Add a test_Automation tool for Creating Multiple Departments
@mcp.tool()
def Manage_Departments (path: str = r"C:\Users\velur\Desktop\Selenium-Python\Termination.xlsx"):
    options = webdriver.ChromeOptions()
    options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    driver = webdriver.Chrome(
    service=ChromeService(ChromeDriverManager().install()), 
    options=options
        
        )
    driver.maximize_window()
    
    #path = "C:/Users/LENOVO/vs_selenium/deptData.xlsx"
    
    rows1 = getRowCount(path,'URL')
    rows2 = getRowCount(path,'DETAILS')
    Base_url = readData(path,"URL",2,1)
    driver.get(Base_url)
    
    UN = readData(path,"LD",2,1)
    PW = readData(path,"LD",2,2)
    driver.find_element(By.ID, "userid").send_keys(UN)
    driver.find_element(By.ID, "password").send_keys(PW)
    driver.find_element(By.ID, "btnActive").click()
    
    # - clicking on home button
    
    xpath_home= '//a[@id="pt1:_UIShome"]/*[name()="svg"]/*[name()="g"][4]/*[name()="path"]'
    HomeClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_home)))
    HomeClick.click()
    
    # - clicking on my client groups
    
    xpath_MCG = '//a[@id="groupNode_workforce_management"]'
    MCGClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_MCG)))
    MCGClick.click()
    
    # - clicking on workforce structure
    
    xpath_WS = '//*[@id="itemNode_workforce_management_workforce_structures_0"]'
    WSClick = WebDriverWait(driver,15).until(EC.element_to_be_clickable((By.XPATH, xpath_WS)))
    WSClick.click()
    
    # - searching for manage department and clicking on it
    
    #xpath_SB = "//input[@placeholder='Search for tasks']"
    #searchBox = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_SB)))
    #searchBox.send_keys("manage department" + Keys.ENTER)
    
    WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Manage Departments")))
    link = driver.find_element(By.PARTIAL_LINK_TEXT, "Manage Departments")
    link.click()
    
    for r in range (2,rows2+1):
        xpath_create = "//a[./span[text()='Create']]"
        createClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_create)))
        createClick.click()
    
        xpath_calender = "//input[@aria-label='Effective Start Date']"
        searchBox = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_calender)))
        searchBox.clear()
        date = readData(path,"DETAILS",r,1)
        date_string = date.strftime('%d-%b-%Y')
        searchBox.send_keys(date_string)
    
        xpath_deptSelect = "//a[@title='Search: Department Set']"
        deptSelectClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_deptSelect)))
        deptSelectClick.click()
    
        WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
        link = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
        link.click()
    
        xpath_RDSN = "//input[@aria-label=' Reference Data Set Name']"
        RDSN_SB = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_RDSN)))
        RDSN_SB.clear()
        RDS = readData(path,"DETAILS",r,2)
        RDSN_SB.send_keys(RDS+ Keys.ENTER)
    
        xpath_selectRDSN = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:AP1:setName2Id_afrLovInternalTableId::db"]/table/tbody/tr/td[1]'
        selectRDSN = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_selectRDSN)))
        selectRDSN.click()
    
        xpath_okClick = "//button[text()='OK']"
        okClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_okClick)))
        okClick.click()
        time.sleep(2)
    
        NM = readData(path,"DETAILS",r,3)
        driver.switch_to.active_element.send_keys(Keys.TAB + NM)
    
        xpath_statusDD = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:AP1:selectOneChoice2::content"]'
        statusDD = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_statusDD)))
        statusDD.click()
    
        xpath_Astatus = "//li[text()='Active']"
   
        # 1. Click 'Active' status (This action triggers the Name Exists validation)
        Astatus = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath_Astatus)))
        Astatus.click()
        time.sleep(2)
 
        # - START DIALOG CHECK BLOCK
        # - We use a try/except to check for the MANDATORY presence of the error dialog button (OK)
        # - If the OK button is found, we assume the error occurred and execute the 'except' block.
   
        xpath_error_ok_button = "//button[text()='OK']"
   
        try:    
       
            error_ok_button = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath_error_ok_button)))
            error_ok_button.click()
            time.sleep(1)
 
            xpath_clickCE = "//a[normalize-space(.)='Cancel']"
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath_clickCE))).click()
            time.sleep(1)
 
            xpath_clickYE_confirmation = "//button[@accesskey='Y']"
            try:
                #- If it appears, clicks it.
                WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, xpath_clickYE_confirmation))).click()
                time.sleep(1)
            except TimeoutException:
                pass
 
            writeData(path,"DETAILS",r,4,"NAME EXISTS")
            time.sleep(3)
            continue
           
        except TimeoutException:
        #- IF DIALOG DID NOT APPEAR
               
            xpath_clickNext = "//a[normalize-space(.)='Next']"
            clickNext = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_clickNext)))
            clickNext.click()
 
            xpath_clickNext2 = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:3:AP2:tt1:next"]/a/span'
            clickNext2 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_clickNext2)))
            clickNext2.click()
 
            xpath_clickSubmit = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:4:ap1:tt1:submit"]/a/span'
            clickSubmit = WebDriverWait(driver,15).until(EC.element_to_be_clickable((By.XPATH, xpath_clickSubmit)))
            clickSubmit.click()
 
            xpath_clickYes = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:4:ap1:tt1:okWarningDialog"]'
            clickYes = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_clickYes)))
            clickYes.click()
 
            xpath_clickOK = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:4:ap1:tt1:okConfirmationDialog"]'
            clickOK = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_clickOK)))
            clickOK.click()
 
            writeData(path,"DETAILS",r,4,"DEPT CREATED")
            time.sleep(2)
 
    
    
    driver.quit()
@mcp.tool()
def Termination_Employee (path: str = 
    r"C:\Users\velur\Desktop\Selenium-Python\Termination.xlsx"):
    """Termination of an Employee"""
    #driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    options = webdriver.ChromeOptions()
    options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    driver = webdriver.Chrome(
    service=ChromeService(ChromeDriverManager().install()), 
    options=options
    )
    try:
        driver.maximize_window()
        rows_Termination = getRowCount(path, 'Termination')
        Base_url = readData(path, "URL", 2, 1)
        driver.get(Base_url)
        
        UN = readData(path, "LD", 2, 1)
        PW = readData(path, "LD", 2, 2)
        
        driver.find_element(By.ID, "userid").send_keys(UN)
        driver.find_element(By.ID, "password").send_keys(PW)
        driver.find_element(By.ID, "btnActive").click()
        
        
        # - clicking on home button
        xpath_home= '//a[@id="pt1:_UIShome"]/*[name()="svg"]/*[name()="g"][4]/*[name()="path"]'
        HomeClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_home)))
        HomeClick.click()
        # - clicking on my client groups
        xpath_MCG = '//a[@id="groupNode_workforce_management"]'
        MCGClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_MCG)))
        MCGClick.click()
        # - clicking on person Managment button
        
        xpath_MCG = '//a[@id="itemNode_workforce_management_person_management_0"]'
        MCGClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_MCG)))
        MCGClick.click()
        
        #Enter person number
        # 🔑 CRITICAL FIX 2: Added +1 to the range to include the last row of data
        created = 0
        for r in range(2, rows_Termination+1):
        
            # Read the data from the 'Termination' sheet, column 1
            excel_person_number = readData(path, 'Termination', r, 1)
        
            xpath_PersonNumber = "//*[@id='_FOpt1:_FOr1:0:_FONSr2:0:MAt1:0:pt1:Perso1:0:SP3:q1:value10::content']"
            
            person_number_input = WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable((By.XPATH, xpath_PersonNumber))
            )
            
            # Clear and send keys for the current person number
            person_number_input.clear()
            person_number_input.send_keys(excel_person_number + Keys.ENTER)
            
            # Wait for the search result before proceeding to the next row
            time.sleep(5) 
            xpath_PersonNumber_Search = "//button[text()='Search']"
        
            PersonSearch = WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable((By.XPATH, xpath_PersonNumber_Search))
            )
            PersonSearch.click()
            print("Mahesh1")
            #click on Action button
            action_icon_xpath = "//button[@title='Actions']"
            action_button = WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable((By.XPATH, action_icon_xpath))
            )
            
            #//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAt1:0:pt1:Perso1:0:SP3:table1:_ATp:table2:0:cil1"]/img
            action_button.click()
            #time.sleep(5)
            print("Mahesh2")
            # Click Person & Employment
            xpath_Person = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAt1:0:pt1:Perso1:0:SP3:table1:am2:dc_i1:3:dcm1"]/td[2]'
            Person_emp = WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable((By.XPATH, xpath_Person))
            )
            Person_emp.click()
            
            print("Mahesh3")
            # Click Work Relationship
            xpath_Workforce = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAt1:0:pt1:Perso1:0:SP3:table1:am2:dc_i1:3:dci1:12:dccmi1"]/td[2]'
            Workforce = WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable((By.XPATH, xpath_Workforce))
            )
            Workforce.click()
            print("Mahesh4")
            # ACTIONS
            WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@title='Actions']"))
                    ).click()
            print("Mahesh10")
        
            time.sleep(5)
        
            # TERMINATION BUTTON
            WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//td[text()='Terminate']"))).click()
            time.sleep(5)
            
            # time.sleep(5) # This was at the end of the original script
            print("Mahesh11")
        
            #Enter Action details from Excel
            excel_Action = readData(path, 'Termination', r, 2)
            xpath_actionDD = "//*[contains(@id, 'Action::content')]"
            actionDD = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, xpath_actionDD)))
            actionDD.send_keys(excel_Action + Keys.ENTER)
            time.sleep(5)
            
            #Notificaton Date
            excel_Notification_Date = readData(path, 'Termination', r, 3)
            xpath_NotificationDate = "//input[@aria-label='Notification Date']"
            NotificationDate = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_NotificationDate)))
            NotificationDate.clear()
            date_string = excel_Notification_Date.strftime('%d-%b-%Y')
            NotificationDate.send_keys(date_string + Keys.ENTER)
        
            #Enter Termination Date 
            excel_Termination_Date = readData(path, 'Termination', r, 4)
            xpath_TerminationDate = "//input[@aria-label='Termination Date']"
            TerminationDate = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_TerminationDate)))
            TerminationDate.clear()
            date_string_Termi = excel_Termination_Date.strftime('%d-%b-%Y')
            TerminationDate.send_keys(date_string_Termi + Keys.ENTER)
            time.sleep(5)
            
            # select Recommand for Rehire
            excel_Rehire = readData(path, 'Termination', r, 5)
            xpath_RehireRecom_Input = "//*[contains(@id, 'RehireRecom::content')]" 
            rehireField = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, xpath_RehireRecom_Input))
            )    
            rehireField.send_keys(excel_Rehire + Keys.ENTER)
            time.sleep(5)
            #Click on Review
            xpath_Review= '//button[text()="Review"]'
            Review = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Review)))
            Review.click()
            print("Mahesh7")
            time.sleep(10)
            #Submit
            xpath_Submit1 = "//span[normalize-space(.)='Submit']"
            Submit = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Submit1)))
            Submit.click()
            time.sleep(5)
            print("submit clicked")
            
            #Click on Yes in Warning Box
            
            #//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAt2:1:r1:0:r1:1:pt1:sp1:tt1:okWarningDialog"]
            #
            xpath_Click_Yes= "//button[@accesskey='Y']"
            Click_Yes = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Click_Yes)))
            Click_Yes.click()
            time.sleep(5)
            
            #
            ##Click on Ok in Warnming Box
            #
            xpath_Click_Ok= "//button[@accesskey='K']"
            Click_Ok = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Click_Ok)))
            Click_Ok.click()
        
            xpath_closeB = "//button[contains(., 'Close')]"
            WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_closeB))).click()
            time.sleep(5)
            created += 1
            return {"status": "completed", "created": created}
    finally:
        time.sleep(2)
        try:
            driver.quit()
        except Exception:
            pass
   

if __name__ == "__main__":
    transport = "sse"
    if transport == "stdio":
        print("Running server with stdio transport")
        mcp.run(transport="stdio")
    elif transport == "sse":
        print("Running server with SSE transport")
        mcp.run(transport="sse")
    elif transport == "streamable-http":
        print("Running server with Streamable HTTP transport")
        mcp.run(transport="streamable-http")
    else:
        raise ValueError(f"Unknown transport: {transport}")
